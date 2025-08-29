
import io, os, csv, json
from typing import Optional
from fastapi import FastAPI, UploadFile, File, Query, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from starlette.staticfiles import StaticFiles
from pydantic import BaseModel
from dotenv import load_dotenv

from .extractors.ocr import extract_text_from_file
from .extractors.heuristics import extract_fields_heuristic
from .extractors.validate import validate_extraction
from .extractors.postprocess import autofill_amounts
from .extractors.llm import extract_fields_llm, llm_available
from .extractors.templates import extract_fields_template

load_dotenv()

app = FastAPI(title="Invoice Extractor", version="0.3.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class ExtractResponse(BaseModel):
    data: dict
    method: str
    validations: dict

@app.get("/api/health")
def health():
    return {"status": "ok"}

@app.post("/api/extract", response_model=ExtractResponse)
async def extract(file: UploadFile = File(...), method: Optional[str] = Query("auto")):
    try:
        content = await file.read()
        text = extract_text_from_file(filename=file.filename, data=content)

        used_method = ""
        result = None

        # 1) Template
        if method in ["template", "auto"]:
            tpl_res = extract_fields_template(text)
            if tpl_res:
                result = tpl_res
                used_method = "template"

        # 2) LLM
        if result is None and (method == "llm" or (method == "auto" and llm_available())):
            try:
                result = extract_fields_llm(text)
                used_method = "llm"
            except Exception:
                result = None

        # If LLM succeeded but left some fields empty, augment with heuristics
        if result is not None and used_method.startswith("llm"):
            try:
                heur = extract_fields_heuristic(text)
                if isinstance(heur, dict):
                    # Merge only missing/empty fields from heuristics output
                    def _is_empty(v):
                        return v is None or (isinstance(v, str) and not v.strip())

                    # Top-level simple fields to backfill
                    for key in [
                        "variabilni_symbol",
                        "datum_vystaveni",
                        "datum_splatnosti",
                        "duzp",
                        "castka_bez_dph",
                        "dph",
                        "castka_s_dph",
                        "platba_zpusob",
                        "banka_prijemce",
                        "ucet_prijemce",
                    ]:
                        if (key in heur) and _is_empty(result.get(key)) and not _is_empty(heur.get(key)):
                            result[key] = heur.get(key)

                    # Currency: also replace if LLM produced an unknown token
                    _known = {"CZK","EUR","USD","GBP","PLN","HUF","CHF","SEK","NOK","DKK","JPY","CNY","AUD","CAD"}
                    cur = (result.get("mena") or "").strip().upper()
                    if cur not in _known and not _is_empty(heur.get("mena")):
                        result["mena"] = heur.get("mena")

                    # Nested supplier dictionary
                    llm_sup = result.get("dodavatel") or {}
                    heur_sup = heur.get("dodavatel") or {}
                    if isinstance(llm_sup, dict) and isinstance(heur_sup, dict):
                        for k in ["nazev", "ico", "dic", "adresa"]:
                            if _is_empty(llm_sup.get(k)) and not _is_empty(heur_sup.get(k)):
                                llm_sup[k] = heur_sup.get(k)
                        result["dodavatel"] = llm_sup
                    used_method = "llm+heuristic"
            except Exception:
                # Never block extraction if augmentation fails
                pass

        # 3) Heuristic fallback
        if result is None:
            result = extract_fields_heuristic(text)
            used_method = "heuristic" if method != "llm" else "heuristic (fallback)"

        # Postprocess: compute any missing related amounts
        if isinstance(result, dict):
            result = autofill_amounts(result)
        else:
            result = {}

        validations = validate_extraction(result)
        return ExtractResponse(data=result, method=used_method, validations=validations)
    except Exception:
        # Never fail hard; always return a safe payload so the frontend can proceed
        return ExtractResponse(data={}, method="error", validations={})

# -------- Export endpoint --------
def _flatten_dict(d, prefix=""):
    rows = []
    for k, v in d.items():
        key = f"{prefix}{k}" if not prefix else f"{prefix}.{k}"
        if isinstance(v, dict):
            rows.extend(_flatten_dict(v, key))
        else:
            rows.append((key, "" if v is None else v))
    return rows

def _to_txt(d: dict) -> bytes:
    rows = _flatten_dict(d)
    buf = io.StringIO()
    for k, v in rows:
        buf.write(f"{k}: {v}\n")
    return buf.getvalue().encode("utf-8")

def _to_csv(d: dict) -> bytes:
    rows = _flatten_dict(d)
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Key", "Value"])
    for k, v in rows:
        w.writerow([k, v])
    return out.getvalue().encode("utf-8")

def _to_xlsx(d: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.append(["Key", "Value"])
    for k, v in _flatten_dict(d):
        ws.append([k, v])
    # autosize
    for col in [1, 2]:
        max_len = 0
        for cell in ws[get_column_letter(col)]:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 80)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()

@app.post("/api/export")
async def export_file(payload: dict = Body(...)):
    fmt = (payload.get("format") or "json").lower()
    data = payload.get("data") or {}
    filename = payload.get("filename") or "invoice_export"
    if fmt == "json":
        content = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
        media = "application/json"
        ext = "json"
    elif fmt == "txt":
        content = _to_txt(data)
        media = "text/plain"
        ext = "txt"
    elif fmt == "csv":
        content = _to_csv(data)
        media = "text/csv"
        ext = "csv"
    elif fmt == "xlsx":
        content = _to_xlsx(data)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"
    else:
        return JSONResponse({"error": "Unsupported format"}, status_code=400)

    headers = {"Content-Disposition": f'attachment; filename="{filename}.{ext}"'}
    return StreamingResponse(io.BytesIO(content), media_type=media, headers=headers)

# --- Serve frontend last ---
from pathlib import Path as _Path
_FRONTEND_DIR = str((_Path(__file__).resolve().parents[1] / "frontend").resolve())
app.mount("/", StaticFiles(directory=_FRONTEND_DIR, html=True), name="frontend")
